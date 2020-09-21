import os
import xlrd
import re
import pyperclip
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from rich import print


def init():
    print('初始化中，请稍等...')
    version = 1.3
    print('此程序版本：' + str(version))
    try:
        urlgithub = 'https://raw.githubusercontent.com/evilbutcher/Python/master/evanescent/release.json'
        update = requests.get(urlgithub)
        ver = update.json()
        if ver['releases'][0]['version'] > version:
            print('[bold yellow]更新[/bold yellow]啦！从GitHub获取更新详情成功！\n最新版本是：' +
                  str(ver['releases'][0]['version']))
            print('更新内容是：' + ver['releases'][0]['details'])
            print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
        else:
            print('检测更新完成，暂无更新\n')
    except (Exception):
        urlgitee = 'https://gitee.com/evilbutcher/Python/raw/master/evanescent/release.json'
        try:
            update = requests.get(urlgitee)
            ver = update.json()
            if ver['releases'][0]['version'] > version:
                print('[bold red]更新[/bold red]啦！从Gitee获取更新详情成功！\n最新版本是：' +
                      str(ver['releases'][0]['version']))
                print('更新内容是：' + ver['releases'][0]['details'])
                print('可前往：https://github.com/evilbutcher/Python 查看Releases\n')
            else:
                print('检测更新完成，暂无更新\n')
        except Exception as e:
            print('检测更新失败，原因：')
            print(str(e) + '\n')
    try:
        path = Path('origindata/')
        if path.is_dir() is False:
            os.mkdir(path)
        a = pyperclip.paste()
        if a == 'debugturnon':
            return True
    except Exception as e:
        print('初始化[bold red]失败[/bold red]，原因：' + str(e))


def dealxlsx(path: str, name: str, canprint: bool):
    try:
        workbook = xlrd.open_workbook(path + '/' + name)
        sheet_name = workbook.sheet_names()[0]
        print('\n读取的Excel表名为：' + sheet_name)
        sheet = workbook.sheet_by_index(0)
        allrows = sheet.nrows
        allcols = sheet.ncols - 1
        print('一共有：' + str(allcols) + '列  ' + str(allrows) + '行（第一列时间未计入统计）')
        secend = input('请输入基准点（第几秒）：')
        valueforall = input('请输入一个基准值（自定即可）：')
        startpoint = int(secend)
        startvalue = sheet.cell(startpoint, 0).value
        print('所选取的修改起始点为：' + str(int(startvalue)) + '秒的数据')
        wb = Workbook()  # 引入result.xlsx工作表
        result = wb["Sheet"]
        if result.cell(1, 1).value is None:  # 因为两个库起始index不同，先设置一下第一行
            result.cell(1, 1).value = 0
            if canprint is True:
                print('\n1列1行 写入时间：0秒\n')
        for col in range(1, allcols + 1):  # 获取数据excel每列与基准值的差值
            minus = sheet.cell(startpoint, col).value - float(valueforall)
            print('第' + str(col) + '列截取的起始数据点为：' +
                  str(sheet.cell(startpoint, col).value))
            print('第' + str(col) + '列与基准值的差为：' + str(minus) + '\n')
            vcols = sheet.col_values(col)
            result.cell(1, col + 1).value = float(
                valueforall)  # 因为两个库起始index不同，先设置一下第一行
            if canprint is True:
                print(
                    str(col + 1) + '列1行' + '写入数据：' + str(float(valueforall)) +
                    '\n')
            for num in range(1, len(vcols)):  # 获取数据excel每行的数据，写入index+1的位置
                if result.cell(num + 1, 1).value is None:
                    result.cell(num + 1, 1).value = int(num)
                    if canprint is True:
                        print('1列' + str(num + 1) + '行 写入时间：' + str(int(num)) +
                              '秒' + '\n')
                if num > startpoint:
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '读取数据：' +
                            str(sheet.cell(num, col).value))
                    if isinstance(sheet.cell(num, col).value, float) is True:
                        result.cell(
                            num + 1,
                            col + 1).value = sheet.cell(num, col).value - minus
                    else:
                        print(
                            str(col) + '列' + str(num + 1) +
                            '行 读取数据出现[bold red]错误[/bold red]，请检查相应位置，此次选用上一个位置的数值')
                        if isinstance(sheet.cell(num - 1, col).value,
                                      float) is True:
                            result.cell(num + 1, col + 1).value = sheet.cell(
                                num - 1, col).value - minus
                        else:
                            print(
                                str(col) + '列' + str(num + 1) +
                                '行 读取数据出现[bold red]连续错误[/bold red]，请检查相应位置')
                            return
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '写入数据：' +
                            str(float(sheet.cell(num, col).value - minus)) +
                            '\n')
                else:
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '读取数据：' +
                            str(sheet.cell(num, col).value))
                    result.cell(num + 1, col + 1).value = float(valueforall)
                    if canprint is True:
                        print(
                            str(col + 1) + '列' + str(num + 1) + '行' + '写入数据：' +
                            str(float(valueforall)) + '\n')
        wb.save("result.xlsx")
        print('[bold green]数据处理完成[/bold green]')
    except Exception as e:
        print('处理excel[bold red]失败[/bold red]，原因：' + str(e))


def setcolor():
    wb = load_workbook(r'result.xlsx')
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]
    rows = ws.max_row
    columns = ws.max_column
    color = 'FF6666'
    fille = PatternFill('solid', fgColor=color)
    num = int(input('请输入本次实验的平行实验次数：'))
    waittocolor = []
    startnum = 1
    for m in range(0, num):
        startnum = startnum + 1
        waittocolor.append(startnum)
    for i in waittocolor:
        for k in range(i, columns + 1, num * 2):
            for j in range(1, rows + 1):
                ws.cell(j, k).fill = fille
    wb.save(r'result.xlsx')
    print('着色区分完成')


def main():
    try:
        canprint = False
        if init() is True:
            canprint = True
        name = input('请输入要处理的文件名')
        if re.search(r'xlsx|xls', name) is None:
            namewithsuffix4 = name + '.xlsx'
            path = Path('origindata/' + namewithsuffix4)
            if path.is_file() is True:
                dealxlsx('origindata/', namewithsuffix4, canprint)
            else:
                namewithsuffix3 = name + '.xls'
                path = Path('origindata/' + namewithsuffix3)
                if path.is_file() is True:
                    dealxlsx('origindata/', namewithsuffix3, canprint)
        else:
            dealxlsx('origindata/', name, canprint)
        exist = Path('origindata/result.xlsx')
        if exist.is_file() is True:
            setcolor()
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')
    except Exception as e:
        print('主函数运行[bold red]出现错误[/bold red]，原因：' + str(e))
        input('如有问题请前往 https://github.com/evilbutcher/Python 提出issue，请按任意键退出')


if __name__ == "__main__":
    main()
